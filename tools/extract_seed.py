"""Convert the supplied CBDS workbook product table into application seed JSON."""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def main(source: str, output: str) -> None:
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    products = []
    for row in range(19, 50):
        if not ws.cell(row, 2).value:
            continue
        products.append({
            "code": str(ws.cell(row, 2).value).strip(),
            "category": str(ws.cell(row, 3).value or "Uncategorized").strip(),
            "name": str(ws.cell(row, 4).value or "Unnamed product").strip(),
            "size": str(ws.cell(row, 6).value or "").strip(),
            "variant": str(ws.cell(row, 7).value or "").strip(),
            "quantity": int(ws.cell(row, 8).value or 0),
            "weight_g": float(ws.cell(row, 9).value or 0),
            "shipping_rate_kg": float(ws.cell(row, 10).value or 0),
            "purchase_rmb": float(ws.cell(row, 11).value or 0),
        })
    Path(output).write_text(json.dumps({"products": products}, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(products)} products to {output}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
