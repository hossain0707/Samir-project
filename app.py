from __future__ import annotations

import hashlib
import json
import os
import secrets
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk

from PIL import Image, ImageTk


APP_NAME = "CBDs"
APP_VERSION = "2.0.0"
DEFAULT_FONT = "{Segoe UI} 10"
BG = "#F4F7FB"
CARD = "#FFFFFF"
INK = "#172033"
MUTED = "#6B7280"
NAV = "#101A2F"
PRIMARY = "#6C4CF1"
PRIMARY_DARK = "#5535DC"
GREEN = "#0FA968"
ORANGE = "#F59E0B"
RED = "#E6465C"
BORDER = "#E5EAF2"


def app_data_dir() -> Path:
    root = Path(os.getenv("LOCALAPPDATA", Path.home())) / "CBDs"
    root.mkdir(parents=True, exist_ok=True)
    return root


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / name


def store_product_image(source: str) -> str:
    """Copy a selected product image into CBDs-managed application storage."""
    if not source:
        return ""
    source_path = Path(source)
    if not source_path.is_file():
        raise ValueError("The selected product image could not be found")
    target_dir = app_data_dir() / "product_images"
    target_dir.mkdir(parents=True, exist_ok=True)
    if source_path.parent == target_dir:
        return str(source_path)
    target = target_dir / f"{datetime.now():%Y%m%d%H%M%S%f}{source_path.suffix.lower()}"
    shutil.copy2(source_path, target)
    return str(target)


def money(value: float, currency: str = "BDT") -> str:
    label = "BDT" if currency == "BDT" else "RMB"
    return f"{label} {value:,.2f}"


def hash_password(password: str, salt: bytes | None = None) -> tuple[str, str]:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 240_000)
    return salt.hex(), digest.hex()


def generate_batch_code(prefix: str, supply_date: str) -> str:
    parsed = datetime.strptime(supply_date, "%Y-%m-%d")
    clean = "".join(ch for ch in prefix.upper() if ch.isalnum())[-4:] or "CBDS"
    return f"{clean}-{parsed:%Y%m%d}"


def generate_sku(weight_g: float, variant: str, retail_bdt: float) -> str:
    weight = max(0, int(round(float(weight_g))))
    letter = next((ch.upper() for ch in variant.strip() if ch.isalpha()), "X")
    price = max(0, int(round(float(retail_bdt))))
    return f"{weight:03d}{letter}{price:03d}"


def calculate_partner_settlement(invested: float, paid_overhead: float, equal_share: float) -> float:
    """Positive means the partner should receive money; negative means money is due."""
    return float(invested) + float(paid_overhead) - float(equal_share)


class Database:
    def __init__(self) -> None:
        self.path = app_data_dir() / "cbds.db"
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._migrate()
        self._seed()

    def _migrate(self) -> None:
        self.conn.executescript("""
        CREATE TABLE IF NOT EXISTS settings(
          key TEXT PRIMARY KEY, value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS users(
          id INTEGER PRIMARY KEY, name TEXT NOT NULL, username TEXT UNIQUE NOT NULL,
          role TEXT NOT NULL CHECK(role IN ('admin','partner')),
          salt TEXT NOT NULL, password_hash TEXT NOT NULL, active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS products(
          id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE NOT NULL,
          category TEXT NOT NULL, name TEXT NOT NULL, size TEXT, variant TEXT,
          quantity INTEGER NOT NULL CHECK(quantity >= 0), weight_g REAL NOT NULL CHECK(weight_g >= 0),
          shipping_rate_kg REAL NOT NULL CHECK(shipping_rate_kg >= 0), purchase_rmb REAL NOT NULL CHECK(purchase_rmb >= 0),
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS expenses(
          id INTEGER PRIMARY KEY AUTOINCREMENT, expense_date TEXT NOT NULL,
          supplier TEXT NOT NULL, category TEXT NOT NULL, description TEXT NOT NULL,
          quantity INTEGER NOT NULL DEFAULT 1, amount_rmb REAL NOT NULL DEFAULT 0,
          paid_by TEXT NOT NULL, voucher_status TEXT NOT NULL DEFAULT 'Pending', notes TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS partners(
          id INTEGER PRIMARY KEY AUTOINCREMENT, partner_code TEXT UNIQUE NOT NULL,
          name TEXT NOT NULL, phone TEXT, email TEXT, active INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS investments(
          id INTEGER PRIMARY KEY AUTOINCREMENT, partner_id INTEGER NOT NULL REFERENCES partners(id),
          amount_bdt REAL NOT NULL CHECK(amount_bdt >= 0), investment_date TEXT NOT NULL,
          payment_method TEXT, reference_no TEXT, notes TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS overheads(
          id INTEGER PRIMARY KEY AUTOINCREMENT, batch_code TEXT NOT NULL, expense_date TEXT NOT NULL,
          category TEXT NOT NULL, description TEXT NOT NULL, amount_bdt REAL NOT NULL CHECK(amount_bdt >= 0),
          paid_by_partner_id INTEGER REFERENCES partners(id), payment_method TEXT, reference_no TEXT,
          voucher TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS customers(
          id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, phone TEXT UNIQUE NOT NULL,
          address TEXT NOT NULL, delivery_status TEXT NOT NULL DEFAULT 'Active', notes TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS suppliers(
          id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, contact_person TEXT,
          phone TEXT, wechat TEXT, address TEXT, quality_rating REAL NOT NULL DEFAULT 0,
          last_price_rmb REAL NOT NULL DEFAULT 0, notes TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS supplier_prices(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          supplier_id INTEGER NOT NULL REFERENCES suppliers(id) ON DELETE CASCADE,
          product_name TEXT NOT NULL,
          quoted_price_rmb REAL NOT NULL CHECK(quoted_price_rmb >= 0),
          quoted_on TEXT NOT NULL, notes TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS batches(
          id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE NOT NULL, supply_date TEXT NOT NULL,
          shipment_rate_kg REAL NOT NULL DEFAULT 0, sourcing_cost_rmb REAL NOT NULL DEFAULT 0,
          status TEXT NOT NULL DEFAULT 'Open', notes TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sales(
          id INTEGER PRIMARY KEY AUTOINCREMENT, invoice_no TEXT UNIQUE NOT NULL,
          customer_id INTEGER REFERENCES customers(id), sale_date TEXT NOT NULL,
          sale_type TEXT NOT NULL CHECK(sale_type IN ('Retail','Wholesale')),
          delivery_status TEXT NOT NULL DEFAULT 'Pending', total_bdt REAL NOT NULL DEFAULT 0,
          paid_bdt REAL NOT NULL DEFAULT 0, notes TEXT, created_by INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sale_items(
          id INTEGER PRIMARY KEY AUTOINCREMENT, sale_id INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
          product_id INTEGER NOT NULL REFERENCES products(id), quantity INTEGER NOT NULL CHECK(quantity > 0),
          unit_price_bdt REAL NOT NULL CHECK(unit_price_bdt >= 0)
        );
        CREATE TABLE IF NOT EXISTS audit_log(
          id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER REFERENCES users(id),
          action TEXT NOT NULL, entity TEXT NOT NULL, entity_id INTEGER, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """)
        self._add_column("products", "image_path", "TEXT")
        self._add_column("products", "batch_code", "TEXT")
        self._add_column("expenses", "batch_code", "TEXT")
        defaults = {
            "rmb_rate": "19.2", "retail_margin": "30", "wholesale_margin": "20",
            "sourcing_cost_rmb": "277.56", "company_prefix": "CBDS", "supply_date": "2026-08-01"
        }
        self.conn.executemany("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", defaults.items())
        self.conn.commit()

    def _add_column(self, table: str, column: str, definition: str) -> None:
        columns = {row[1] for row in self.conn.execute(f"PRAGMA table_info({table})")}
        if column not in columns:
            self.conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    def _seed(self) -> None:
        if not self.conn.execute("SELECT 1 FROM users LIMIT 1").fetchone():
            salt, digest = hash_password("ChangeMe123!")
            self.conn.execute(
                "INSERT INTO users(name,username,role,salt,password_hash) VALUES(?,?,?,?,?)",
                ("Main Administrator", "admin", "admin", salt, digest),
            )
        if not self.conn.execute("SELECT 1 FROM products LIMIT 1").fetchone():
            path = resource_path("sample_data.json")
            if path.exists():
                rows = json.loads(path.read_text(encoding="utf-8"))["products"]
                self.conn.executemany(
                    "INSERT INTO products(code,category,name,size,variant,quantity,weight_g,shipping_rate_kg,purchase_rmb) VALUES(:code,:category,:name,:size,:variant,:quantity,:weight_g,:shipping_rate_kg,:purchase_rmb)", rows
                )
        if not self.conn.execute("SELECT 1 FROM expenses LIMIT 1").fetchone():
            expenses = [
                ("2026-08-11", "Beijing to Baigo", "Transportation", "High-speed road fees", 2, 77.56, "Flower", "Submitted"),
                ("2026-08-01", "Warehouse", "Logistics", "Estimated warehouse delivery", 1, 200, "Flower", "Pending"),
            ]
            self.conn.executemany("INSERT INTO expenses(expense_date,supplier,category,description,quantity,amount_rmb,paid_by,voucher_status) VALUES(?,?,?,?,?,?,?,?)", expenses)
        if not self.conn.execute("SELECT 1 FROM partners LIMIT 1").fetchone():
            self.conn.executemany("INSERT INTO partners(partner_code,name) VALUES(?,?)", [("P01","Samir"),("P02","Flower"),("P03","Mustafa")])
        active_code = generate_batch_code(self.setting("company_prefix"), self.setting("supply_date"))
        self.conn.execute("INSERT OR IGNORE INTO batches(code,supply_date,shipment_rate_kg,sourcing_cost_rmb) VALUES(?,?,?,?)", (active_code,self.setting("supply_date"),780,float(self.setting("sourcing_cost_rmb"))))
        self.conn.commit()

    def setting(self, key: str) -> str:
        row = self.conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row[0] if row else ""

    def set_settings(self, values: dict[str, str]) -> None:
        self.conn.executemany("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", values.items())
        self.conn.commit()

    def authenticate(self, username: str, password: str):
        row = self.conn.execute("SELECT * FROM users WHERE lower(username)=lower(?) AND active=1", (username.strip(),)).fetchone()
        if not row:
            return None
        _, digest = hash_password(password, bytes.fromhex(row["salt"]))
        return row if secrets.compare_digest(digest, row["password_hash"]) else None

    def products(self, query: str = ""):
        term = f"%{query.strip()}%"
        return self.conn.execute("SELECT * FROM products WHERE code LIKE ? OR name LIKE ? OR category LIKE ? OR variant LIKE ? ORDER BY id", (term, term, term, term)).fetchall()

    def expenses(self):
        return self.conn.execute("SELECT * FROM expenses ORDER BY expense_date DESC,id DESC").fetchall()

    def save_product(self, values: tuple, product_id: int | None = None) -> None:
        if product_id:
            self.conn.execute("UPDATE products SET code=?,category=?,name=?,size=?,variant=?,quantity=?,weight_g=?,shipping_rate_kg=?,purchase_rmb=?,image_path=?,batch_code=? WHERE id=?", values + (product_id,))
        else:
            self.conn.execute("INSERT INTO products(code,category,name,size,variant,quantity,weight_g,shipping_rate_kg,purchase_rmb,image_path,batch_code) VALUES(?,?,?,?,?,?,?,?,?,?,?)", values)
        self.conn.commit()

    def delete_product(self, product_id: int) -> None:
        self.conn.execute("DELETE FROM products WHERE id=?", (product_id,))
        self.conn.commit()

    def save_expense(self, values: tuple) -> None:
        self.conn.execute("INSERT INTO expenses(expense_date,supplier,category,description,quantity,amount_rmb,paid_by,voucher_status,notes,batch_code) VALUES(?,?,?,?,?,?,?,?,?,?)", values)
        self.conn.commit()

    def log(self, user_id: int | None, action: str, entity: str, entity_id: int | None = None) -> None:
        self.conn.execute("INSERT INTO audit_log(user_id,action,entity,entity_id) VALUES(?,?,?,?)", (user_id,action,entity,entity_id)); self.conn.commit()

    def partners(self):
        return self.conn.execute("SELECT * FROM partners WHERE active=1 ORDER BY id").fetchall()

    def partner_summary(self):
        overhead_total = float(self.conn.execute("SELECT COALESCE(SUM(amount_bdt),0) FROM overheads").fetchone()[0])
        partners = self.partners(); share = overhead_total / len(partners) if partners else 0
        result=[]
        for p in partners:
            invested=float(self.conn.execute("SELECT COALESCE(SUM(amount_bdt),0) FROM investments WHERE partner_id=?",(p["id"],)).fetchone()[0])
            paid=float(self.conn.execute("SELECT COALESCE(SUM(amount_bdt),0) FROM overheads WHERE paid_by_partner_id=?",(p["id"],)).fetchone()[0])
            result.append((p,invested,paid,share,calculate_partner_settlement(invested,paid,share)))
        return result

    def create_sale(self, customer_id: int, product_id: int, quantity: int, sale_type: str, unit_price: float, delivery_status: str, user_id: int | None) -> int:
        product=self.conn.execute("SELECT quantity FROM products WHERE id=?",(product_id,)).fetchone()
        if not product or quantity <= 0 or product["quantity"] < quantity: raise ValueError("Insufficient product stock")
        invoice_no=f"INV-{datetime.now():%Y%m%d-%H%M%S}-{secrets.randbelow(900)+100}"
        cur=self.conn.execute("INSERT INTO sales(invoice_no,customer_id,sale_date,sale_type,delivery_status,total_bdt,created_by) VALUES(?,?,?,?,?,?,?)",(invoice_no,customer_id,date.today().isoformat(),sale_type,delivery_status,quantity*unit_price,user_id))
        sale_id=cur.lastrowid
        self.conn.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_bdt) VALUES(?,?,?,?)",(sale_id,product_id,quantity,unit_price))
        self.conn.execute("UPDATE products SET quantity=quantity-? WHERE id=?",(quantity,product_id));self.conn.commit();return sale_id

    def backup(self, destination: str) -> None:
        self.conn.commit(); target=sqlite3.connect(destination)
        with target:self.conn.backup(target)
        target.close()


@dataclass(frozen=True)
class Costing:
    purchase_rmb: float
    shipping_rmb: float
    sourcing_rmb: float
    total_rmb: float
    total_bdt: float
    retail_bdt: float
    wholesale_bdt: float


def calculate_cost(product, total_purchase_rmb: float, sourcing_rmb: float, rate: float, retail: float, wholesale: float) -> Costing:
    purchase = float(product["purchase_rmb"])
    shipping = float(product["weight_g"]) / 1000 * float(product["shipping_rate_kg"])
    allocated = (purchase / total_purchase_rmb * sourcing_rmb) if total_purchase_rmb else 0
    total = purchase + shipping + allocated
    return Costing(purchase, shipping, allocated, total, total * rate, total * rate * (1 + retail / 100), total * rate * (1 + wholesale / 100))


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.db = Database()
        self.user = None
        self.title(APP_NAME)
        self.geometry("1440x880")
        self.minsize(1120, 720)
        self.configure(bg=BG)
        try:
            self.iconbitmap(resource_path("assets/cbds.ico"))
        except tk.TclError:
            pass
        # Braces keep the multi-word Windows font family together. Without
        # them Tk parses "UI" as the point size and raises a TclError.
        self.option_add("*Font", DEFAULT_FONT)
        self._styles()
        self.show_login()

    def _styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("Treeview", background=CARD, fieldbackground=CARD, foreground=INK, rowheight=38, borderwidth=0)
        s.configure("Treeview.Heading", background="#EEF1F7", foreground=MUTED, font=("Segoe UI Semibold", 9), relief="flat", padding=8)
        s.map("Treeview", background=[("selected", "#EDE9FE")], foreground=[("selected", INK)])
        s.configure("TEntry", padding=8, fieldbackground=CARD)
        s.configure("TCombobox", padding=7, fieldbackground=CARD)

    def logo_image(self, size: tuple[int, int]):
        image = Image.open(resource_path("assets/cbds_logo.png")).convert("RGBA")
        image.thumbnail(size, Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(image)

    def clear(self):
        for child in self.winfo_children():
            child.destroy()

    def show_login(self):
        self.clear()
        shell = tk.Frame(self, bg=BG)
        shell.pack(fill="both", expand=True)
        left = tk.Frame(shell, bg=NAV, width=560)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        self.login_logo = self.logo_image((315, 155))
        tk.Label(left, image=self.login_logo, bg="white", bd=0).pack(anchor="w", padx=64, pady=(70, 0))
        tk.Label(left, text="Business clarity,\nfrom source to sale.", fg="white", bg=NAV, justify="left", font=("Segoe UI Semibold", 33)).pack(anchor="w", padx=64, pady=(90, 25))
        tk.Label(left, text="Inventory • Import costing • Partner ledger\nBuilt for China BD Store", fg="#AEB8CC", bg=NAV, justify="left", font=("Segoe UI", 12)).pack(anchor="w", padx=64)
        panel = tk.Frame(shell, bg=BG)
        panel.pack(side="left", fill="both", expand=True)
        form = tk.Frame(panel, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        form.place(relx=.5, rely=.5, anchor="center", width=440, height=500)
        tk.Label(form, text="Welcome back", fg=INK, bg=CARD, font=("Segoe UI Semibold", 25)).pack(anchor="w", padx=50, pady=(50, 8))
        tk.Label(form, text="Sign in to continue to CBDs Workspace", fg=MUTED, bg=CARD).pack(anchor="w", padx=50, pady=(0, 34))
        tk.Label(form, text="Username", fg=INK, bg=CARD, font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=50)
        user = ttk.Entry(form)
        user.pack(fill="x", padx=50, pady=(7, 19))
        user.insert(0, "admin")
        tk.Label(form, text="Password", fg=INK, bg=CARD, font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=50)
        pwd_row = tk.Frame(form, bg=CARD)
        pwd_row.pack(fill="x", padx=50, pady=(7, 11))
        pwd = ttk.Entry(pwd_row, show="●")
        pwd.pack(side="left", fill="x", expand=True)
        show = tk.BooleanVar()
        tk.Checkbutton(pwd_row, text="Show", variable=show, command=lambda: pwd.configure(show="" if show.get() else "●"), bg=CARD, fg=MUTED, activebackground=CARD, bd=0).pack(side="left", padx=(8, 0))
        remember = tk.BooleanVar(value=True)
        tk.Checkbutton(form, text="Remember my username", variable=remember, bg=CARD, fg=MUTED, activebackground=CARD, bd=0).pack(anchor="w", padx=46)
        status = tk.Label(form, text="", fg=RED, bg=CARD)
        status.pack(pady=(6, 0))
        def login(*_):
            row = self.db.authenticate(user.get(), pwd.get())
            if not row:
                status.configure(text="Incorrect username or password")
                return
            self.user = row
            self.show_workspace()
        self.action_button(form, "Sign in securely", login).pack(fill="x", padx=50, pady=(12, 12))
        tk.Label(form, text="First sign-in: admin / ChangeMe123!", fg=MUTED, bg=CARD, font=("Segoe UI", 9)).pack()
        pwd.bind("<Return>", login)
        pwd.focus_set()

    def show_workspace(self):
        self.clear()
        sidebar = tk.Frame(self, bg=NAV, width=236)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        self.sidebar_logo = self.logo_image((174, 88))
        tk.Label(sidebar, image=self.sidebar_logo, bg="white", bd=0).pack(anchor="w", padx=27, pady=(24, 7))
        tk.Label(sidebar, text="CBDs", fg="#8290AA", bg=NAV, font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=27, pady=(0, 24))
        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)
        links = [("▦", "Dashboard", self.dashboard), ("▣", "Inventory", self.inventory), ("↗", "Purchasing", self.expenses), ("◎", "Partners", self.partner_ledger), ("●", "Customers", self.customers), ("◆", "Suppliers", self.suppliers), ("▤", "Sales & invoices", self.sales), ("◫", "Reports", self.reports)]
        if self.user["role"] == "admin": links.append(("⚙", "Settings", self.settings))
        self.nav_buttons = []
        for icon, label, command in links:
            b = tk.Button(sidebar, text=f"  {icon}   {label}", command=lambda c=command, x=label: self.navigate(c, x), bg=NAV, fg="#C7D0E0", activebackground="#1F2A44", activeforeground="white", relief="flat", bd=0, anchor="w", padx=20, pady=9, font=("Segoe UI Semibold", 9), cursor="hand2")
            b.pack(fill="x", padx=10, pady=1)
            self.nav_buttons.append((label, b))
        tk.Frame(sidebar, bg=NAV).pack(fill="both", expand=True)
        tk.Label(sidebar, text=self.user["name"], fg="white", bg=NAV, font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=27)
        tk.Label(sidebar, text=self.user["role"].title(), fg="#8290AA", bg=NAV, font=("Segoe UI", 9)).pack(anchor="w", padx=27, pady=(2, 8))
        tk.Button(sidebar, text="Change password", command=self.password_dialog, fg="#B9A7FF", bg=NAV, activebackground=NAV, relief="flat", bd=0, anchor="w", padx=23).pack(fill="x")
        tk.Button(sidebar, text="Sign out", command=self.show_login, fg="#B9A7FF", bg=NAV, activebackground=NAV, relief="flat", bd=0, anchor="w", padx=23).pack(fill="x", pady=(0, 24))
        self.navigate(self.dashboard, "Dashboard")

    def navigate(self, command, name):
        for label, b in self.nav_buttons:
            b.configure(bg="#241F4F" if label == name else NAV, fg="white" if label == name else "#C7D0E0")
        command()

    def page(self, title: str, subtitle: str):
        for w in self.content.winfo_children():
            w.destroy()
        header = tk.Frame(self.content, bg=BG)
        header.pack(fill="x", padx=34, pady=(27, 20))
        tk.Label(header, text=title, bg=BG, fg=INK, font=("Segoe UI Semibold", 24)).pack(anchor="w")
        tk.Label(header, text=subtitle, bg=BG, fg=MUTED).pack(anchor="w", pady=(4, 0))
        body = tk.Frame(self.content, bg=BG)
        body.pack(fill="both", expand=True, padx=34, pady=(0, 30))
        return body

    def action_button(self, parent, text, command, secondary=False):
        return tk.Button(parent, text=text, command=command, bg=CARD if secondary else PRIMARY, fg=INK if secondary else "white", activebackground="#F7F7FB" if secondary else PRIMARY_DARK, activeforeground=INK if secondary else "white", relief="flat", bd=0, padx=18, pady=10, cursor="hand2", font=("Segoe UI Semibold", 10), highlightthickness=1 if secondary else 0, highlightbackground=BORDER)

    def card(self, parent):
        return tk.Frame(parent, bg=CARD, highlightbackground=BORDER, highlightthickness=1)

    def is_admin(self):
        return bool(self.user and self.user["role"] == "admin")

    def context(self):
        products = self.db.products()
        rate = float(self.db.setting("rmb_rate"))
        retail = float(self.db.setting("retail_margin"))
        wholesale = float(self.db.setting("wholesale_margin"))
        sourcing = float(self.db.setting("sourcing_cost_rmb"))
        purchase = sum(float(p["purchase_rmb"]) for p in products)
        costs = [calculate_cost(p, purchase, sourcing, rate, retail, wholesale) for p in products]
        return products, costs, rate, retail, wholesale, sourcing

    def dashboard(self):
        body = self.page("Good day, Hossain", "Here is the current sourcing and inventory position.")
        products, costs, rate, _, _, _ = self.context()
        metrics = [
            ("Inventory units", str(sum(p["quantity"] for p in products)), "Across all active products", PRIMARY),
            ("Landed inventory cost", money(sum(c.total_bdt * p["quantity"] for c, p in zip(costs, products))), f"At {rate:g} BDT / RMB", GREEN),
            ("Projected retail value", money(sum(c.retail_bdt * p["quantity"] for c, p in zip(costs, products))), "Current retail margin", ORANGE),
            ("Low / out of stock", str(sum(1 for p in products if p["quantity"] <= 2)), "Needs attention", RED),
        ]
        row = tk.Frame(body, bg=BG)
        row.pack(fill="x")
        for label, value, hint, color in metrics:
            c = self.card(row); c.pack(side="left", fill="x", expand=True, padx=(0, 14))
            tk.Frame(c, bg=color, height=4).pack(fill="x")
            tk.Label(c, text=label, bg=CARD, fg=MUTED, font=("Segoe UI Semibold", 9)).pack(anchor="w", padx=20, pady=(18, 7))
            tk.Label(c, text=value, bg=CARD, fg=INK, font=("Segoe UI Semibold", 20)).pack(anchor="w", padx=20)
            tk.Label(c, text=hint, bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(5, 18))
        lower = tk.Frame(body, bg=BG); lower.pack(fill="both", expand=True, pady=(22, 0))
        table_card = self.card(lower); table_card.pack(side="left", fill="both", expand=True, padx=(0, 18))
        tk.Label(table_card, text="Inventory highlights", bg=CARD, fg=INK, font=("Segoe UI Semibold", 14)).pack(anchor="w", padx=20, pady=(18, 3))
        tk.Label(table_card, text="Highest projected retail value", bg=CARD, fg=MUTED).pack(anchor="w", padx=20, pady=(0, 13))
        tree = ttk.Treeview(table_card, columns=("name","stock","cost","retail"), show="headings", height=10)
        for col, label, width in [("name","Product",260),("stock","Stock",70),("cost","Landed",110),("retail","Retail",110)]: tree.heading(col,text=label); tree.column(col,width=width,anchor="e" if col != "name" else "w")
        ranked = sorted(zip(products,costs), key=lambda x:x[1].retail_bdt*x[0]["quantity"], reverse=True)[:10]
        for p,cost in ranked: tree.insert("", "end", values=(p["name"],p["quantity"],money(cost.total_bdt),money(cost.retail_bdt)))
        tree.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        side = self.card(lower); side.pack(side="left", fill="y", ipadx=8)
        tk.Label(side, text="Supply batch", bg=CARD, fg=INK, font=("Segoe UI Semibold", 14)).pack(anchor="w", padx=20, pady=(18, 20))
        prefix = self.db.setting("company_prefix")[-4:].upper(); supply = self.db.setting("supply_date")
        for label,value in [("Supply code", f"{prefix}-{supply.replace('-','')}"),("Supply date",supply),("Exchange rate",f"1 RMB = {rate:g} BDT"),("Active SKUs",str(len(products)))]:
            tk.Label(side,text=label,bg=CARD,fg=MUTED,font=("Segoe UI",9)).pack(anchor="w",padx=20)
            tk.Label(side,text=value,bg=CARD,fg=INK,font=("Segoe UI Semibold",12)).pack(anchor="w",padx=20,pady=(2,17))

    def inventory(self):
        body = self.page("Inventory", "Manage products, stock, shipment inputs, and landed cost.")
        toolbar = tk.Frame(body,bg=BG); toolbar.pack(fill="x",pady=(0,12))
        search = ttk.Entry(toolbar); search.pack(side="left",fill="x",expand=True); search.insert(0,"")
        self.action_button(toolbar,"Search",lambda:self._fill_inventory(tree,search.get()),secondary=True).pack(side="left",padx=8)
        self.action_button(toolbar,"+ Add product",lambda:self.product_dialog()).pack(side="right")
        wrap=self.card(body); wrap.pack(fill="both",expand=True)
        cols=("id","code","name","category","variant","qty","cost","retail")
        tree=ttk.Treeview(wrap,columns=cols,show="headings")
        labels=("ID","Code","Product","Category","Variant","Stock","Landed BDT","Retail BDT")
        for col,label in zip(cols,labels): tree.heading(col,text=label)
        widths=(55,110,230,100,110,70,115,115)
        for col,w in zip(cols,widths): tree.column(col,width=w,anchor="e" if col in ("qty","cost","retail") else "w")
        tree.pack(fill="both",expand=True,padx=15,pady=(15,8))
        actions=tk.Frame(wrap,bg=CARD);actions.pack(fill="x",padx=15,pady=(0,15))
        def selected_id():
            selected=tree.selection(); return int(tree.item(selected[0],"values")[0]) if selected else None
        self.action_button(actions,"Edit selected",lambda:self.product_dialog(selected_id()),secondary=True).pack(side="left")
        def delete():
            pid=selected_id()
            if not self.is_admin(): return messagebox.showwarning("Permission denied","Only an administrator can delete products.")
            if pid and messagebox.askyesno("Delete product","Delete the selected product? This cannot be undone."):
                self.db.delete_product(pid); self._fill_inventory(tree,search.get())
        self.action_button(actions,"Delete",delete,secondary=True).pack(side="left",padx=8)
        search.bind("<Return>",lambda _:self._fill_inventory(tree,search.get()))
        self._fill_inventory(tree,"")

    def _fill_inventory(self, tree, query):
        tree.delete(*tree.get_children())
        products, _, rate, retail, wholesale, sourcing = self.context()
        filtered=self.db.products(query); total=sum(float(p["purchase_rmb"]) for p in products)
        for p in filtered:
            c=calculate_cost(p,total,sourcing,rate,retail,wholesale)
            tag="out" if p["quantity"]==0 else "low" if p["quantity"]<=2 else "normal"
            tree.insert("","end",values=(p["id"],p["code"],p["name"],p["category"],p["variant"],p["quantity"],f"{c.total_bdt:,.2f}",f"{c.retail_bdt:,.2f}"),tags=(tag,))
        tree.tag_configure("out",background="#FDE8EC",foreground=RED);tree.tag_configure("low",background="#FFF7E0",foreground="#9A6700")

    def product_dialog(self, product_id=None):
        row=self.db.conn.execute("SELECT * FROM products WHERE id=?",(product_id,)).fetchone() if product_id else None
        win=tk.Toplevel(self);win.title("Edit product" if row else "Add product");win.geometry("580x790");win.configure(bg=CARD);win.transient(self);win.grab_set()
        tk.Label(win,text="Edit product" if row else "New inventory product",bg=CARD,fg=INK,font=("Segoe UI Semibold",20)).pack(anchor="w",padx=32,pady=(25,18))
        form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=32)
        fields=[("Product code / SKU","code"),("Category","category"),("Product name","name"),("Size","size"),("Variant / color","variant"),("Quantity","quantity"),("Weight (grams)","weight_g"),("Shipment rate / kg (RMB)","shipping_rate_kg"),("Purchase price / unit (RMB)","purchase_rmb"),("Image file path","image_path"),("Batch code","batch_code")]
        entries={}
        for label,key in fields:
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w")
            e=ttk.Entry(form);e.pack(fill="x",pady=(4,11));entries[key]=e
            if row:e.insert(0,row[key] if row[key] is not None else "")
        if not row: entries["batch_code"].insert(0,generate_batch_code(self.db.setting("company_prefix"),self.db.setting("supply_date")))
        def choose_image():
            selected=filedialog.askopenfilename(parent=win,filetypes=[("Product images","*.png *.jpg *.jpeg *.webp"),("All files","*.*")])
            if selected:entries["image_path"].delete(0,"end");entries["image_path"].insert(0,selected)
        def suggest_sku():
            try:
                weight=float(entries["weight_g"].get()); purchase=float(entries["purchase_rmb"].get()); ship=weight/1000*float(entries["shipping_rate_kg"].get())
                retail=(purchase+ship)*float(self.db.setting("rmb_rate"))*(1+float(self.db.setting("retail_margin"))/100)
                entries["code"].delete(0,"end");entries["code"].insert(0,generate_sku(weight,entries["variant"].get(),retail))
            except ValueError:messagebox.showwarning("Missing inputs","Enter weight, variant, shipment rate, and purchase price first.",parent=win)
        quick=tk.Frame(form,bg=CARD);quick.pack(fill="x",pady=(0,8));self.action_button(quick,"Choose image",choose_image,secondary=True).pack(side="left");self.action_button(quick,"Generate SKU",suggest_sku,secondary=True).pack(side="left",padx=8)
        def save():
            try:
                image_path=store_product_image(entries["image_path"].get().strip())
                values=(entries["code"].get().strip(),entries["category"].get().strip(),entries["name"].get().strip(),entries["size"].get().strip(),entries["variant"].get().strip(),int(entries["quantity"].get()),float(entries["weight_g"].get()),float(entries["shipping_rate_kg"].get()),float(entries["purchase_rmb"].get()),image_path,entries["batch_code"].get().strip())
                if not all(values[:3]): raise ValueError("Code, category, and name are required")
                self.db.save_product(values,product_id);win.destroy();self.inventory()
            except (ValueError,sqlite3.IntegrityError) as exc: messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save product",save).pack(fill="x",pady=(5,20))

    def expenses(self):
        body=self.page("Purchasing & expenses","Record sourcing, transport, customs, and partner-paid costs.")
        top=tk.Frame(body,bg=BG);top.pack(fill="x",pady=(0,12));self.action_button(top,"+ Record expense",self.expense_dialog).pack(side="right");self.action_button(top,"+ New batch",self.batch_dialog,secondary=True).pack(side="right",padx=(0,8))
        wrap=self.card(body);wrap.pack(fill="both",expand=True)
        cols=("id","date","batch","supplier","category","description","rmb","bdt","payer","voucher")
        tree=ttk.Treeview(wrap,columns=cols,show="headings")
        for c,l in zip(cols,("ID","Date","Batch","Supplier","Category","Description","RMB","BDT","Paid by","Voucher")):tree.heading(c,text=l)
        for c,w in zip(cols,(45,95,125,140,110,220,85,105,90,90)):tree.column(c,width=w,anchor="e" if c in ("rmb","bdt") else "w")
        rate=float(self.db.setting("rmb_rate"))
        for e in self.db.expenses():tree.insert("","end",values=(e["id"],e["expense_date"],e["batch_code"] or "",e["supplier"],e["category"],e["description"],f'{e["amount_rmb"]:,.2f}',f'{e["amount_rmb"]*rate:,.2f}',e["paid_by"],e["voucher_status"]))
        tree.pack(fill="both",expand=True,padx=15,pady=15)

    def batch_dialog(self):
        win=tk.Toplevel(self);win.title("Create shipment batch");win.geometry("520x560");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="Create shipment batch",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30);entries={}
        defaults=(("Supply date",date.today().isoformat()),("Shipment rate / kg (RMB)","0"),("Sourcing cost (RMB)","0"),("Status","Open"),("Notes",""))
        for label,default in defaults:tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form);e.insert(0,default);e.pack(fill="x",pady=(4,11));entries[label]=e
        def save():
            try:
                supply=entries["Supply date"].get();datetime.strptime(supply,"%Y-%m-%d");rate=float(entries["Shipment rate / kg (RMB)"].get());sourcing=float(entries["Sourcing cost (RMB)"].get());code=generate_batch_code(self.db.setting("company_prefix"),supply)
                if rate<0 or sourcing<0:raise ValueError("Batch costs cannot be negative")
                self.db.conn.execute("INSERT INTO batches(code,supply_date,shipment_rate_kg,sourcing_cost_rmb,status,notes) VALUES(?,?,?,?,?,?)",(code,supply,rate,sourcing,entries["Status"].get().strip() or "Open",entries["Notes"].get().strip()));self.db.conn.commit();self.db.set_settings({"supply_date":supply,"sourcing_cost_rmb":str(sourcing)});win.destroy();messagebox.showinfo("Batch created",f"Active batch: {code}");self.expenses()
            except (ValueError,sqlite3.IntegrityError) as exc:messagebox.showerror("Cannot create batch",str(exc),parent=win)
        self.action_button(form,"Create and activate batch",save).pack(fill="x",pady=10)

    def expense_dialog(self):
        win=tk.Toplevel(self);win.title("Record expense");win.geometry("560x730");win.configure(bg=CARD);win.transient(self);win.grab_set()
        tk.Label(win,text="Record purchasing expense",bg=CARD,fg=INK,font=("Segoe UI Semibold",20)).pack(anchor="w",padx=32,pady=(25,18))
        form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=32)
        specs=[("Expense date","date",date.today().isoformat()),("Batch code","batch_code",generate_batch_code(self.db.setting("company_prefix"),self.db.setting("supply_date"))),("Supplier","supplier",""),("Category","category","Product Sourcing"),("Description","description",""),("Quantity","quantity","1"),("Amount (RMB)","amount","0"),("Paid by","paid_by","Samir"),("Voucher status","voucher","Pending"),("Notes","notes","")]
        entries={}
        for label,key,default in specs:
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w")
            if key in ("category","voucher"):
                values=("Product Sourcing","Transportation","Local Conveyance","Customs Fees","Logistics","Other") if key=="category" else ("Pending","Submitted")
                e=ttk.Combobox(form,values=values,state="readonly");e.set(default)
            else:e=ttk.Entry(form);e.insert(0,default)
            e.pack(fill="x",pady=(4,10));entries[key]=e
        def save():
            try:
                datetime.strptime(entries["date"].get(),"%Y-%m-%d")
                vals=(entries["date"].get(),entries["supplier"].get().strip(),entries["category"].get(),entries["description"].get().strip(),int(entries["quantity"].get()),float(entries["amount"].get()),entries["paid_by"].get().strip(),entries["voucher"].get(),entries["notes"].get().strip(),entries["batch_code"].get().strip())
                if not vals[1] or not vals[3] or not vals[6]:raise ValueError("Supplier, description, and payer are required")
                self.db.save_expense(vals);win.destroy();self.expenses()
            except ValueError as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save expense",save).pack(fill="x",pady=(4,20))

    def partner_ledger(self):
        body=self.page("Partner settlement","Track investments, shared overheads, reimbursements, and partner balances.")
        top=tk.Frame(body,bg=BG);top.pack(fill="x",pady=(0,12))
        self.action_button(top,"+ Record overhead",self.overhead_dialog,secondary=True).pack(side="right")
        self.action_button(top,"+ Record investment",self.investment_dialog).pack(side="right",padx=(0,8))
        if self.is_admin():self.action_button(top,"+ Add partner",self.partner_dialog,secondary=True).pack(side="left")
        wrap=self.card(body);wrap.pack(fill="both",expand=True)
        cols=("code","name","invested","paid","share","balance")
        tree=ttk.Treeview(wrap,columns=cols,show="headings")
        for c,l in zip(cols,("Partner ID","Partner","Investment BDT","Overhead paid","Equal share","Settlement balance")):tree.heading(c,text=l)
        for c,w in zip(cols,(110,180,150,150,140,170)):tree.column(c,width=w,anchor="e" if c not in ("code","name") else "w")
        for p,invested,paid,share,balance in self.db.partner_summary():
            tree.insert("","end",values=(p["partner_code"],p["name"],f"{invested:,.2f}",f"{paid:,.2f}",f"{share:,.2f}",f"{balance:,.2f}"),tags=("receive" if balance>=0 else "due",))
        tree.tag_configure("receive",foreground=GREEN);tree.tag_configure("due",foreground=RED);tree.pack(fill="both",expand=True,padx=15,pady=15)
        tk.Label(wrap,text="Positive balance = amount receivable • Negative balance = amount due",bg=CARD,fg=MUTED).pack(anchor="w",padx=15,pady=(0,15))

    def partner_dialog(self):
        win=tk.Toplevel(self);win.title("Add partner");win.geometry("500x470");win.configure(bg=CARD);win.transient(self);win.grab_set()
        tk.Label(win,text="Add business partner",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30);entries={}
        for label in ("Partner ID","Partner name","Phone","Email"):
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form);e.pack(fill="x",pady=(4,11));entries[label]=e
        def save():
            try:
                code=entries["Partner ID"].get().strip().upper();name=entries["Partner name"].get().strip()
                if not code or not name:raise ValueError("Partner ID and partner name are required")
                self.db.conn.execute("INSERT INTO partners(partner_code,name,phone,email) VALUES(?,?,?,?)",(code,name,entries["Phone"].get().strip(),entries["Email"].get().strip()));self.db.conn.commit();win.destroy();self.partner_ledger()
            except (ValueError,sqlite3.IntegrityError) as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save partner",save).pack(fill="x",pady=10)

    def investment_dialog(self):
        partners=self.db.partners()
        if not partners:return messagebox.showwarning("Partners required","Add at least one active partner first.")
        win=tk.Toplevel(self);win.title("Record investment");win.geometry("520x470");win.configure(bg=CARD);win.transient(self);win.grab_set()
        tk.Label(win,text="Record partner investment",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30)
        choices={f'{p["partner_code"]} — {p["name"]}':p["id"] for p in partners};partner=ttk.Combobox(form,values=list(choices),state="readonly");partner.set(next(iter(choices)))
        specs=[("Partner",partner),("Amount (BDT)",ttk.Entry(form)),("Investment date",ttk.Entry(form)),("Payment method",ttk.Combobox(form,values=("Cash","Bank transfer","bKash","Nagad","Other"))), ("Reference number",ttk.Entry(form)),("Notes",ttk.Entry(form))]
        specs[1][1].insert(0,"0");specs[2][1].insert(0,date.today().isoformat());specs[3][1].set("Bank transfer")
        for label,widget in specs:tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");widget.pack(fill="x",pady=(4,10))
        def save():
            try:
                amount=float(specs[1][1].get());datetime.strptime(specs[2][1].get(),"%Y-%m-%d")
                if amount<=0:raise ValueError("Investment must be greater than zero")
                self.db.conn.execute("INSERT INTO investments(partner_id,amount_bdt,investment_date,payment_method,reference_no,notes) VALUES(?,?,?,?,?,?)",(choices[partner.get()],amount,specs[2][1].get(),specs[3][1].get(),specs[4][1].get().strip(),specs[5][1].get().strip()));self.db.conn.commit();win.destroy();self.partner_ledger()
            except ValueError as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save investment",save).pack(fill="x",pady=8)

    def overhead_dialog(self):
        partners=self.db.partners();win=tk.Toplevel(self);win.title("Record overhead");win.geometry("540x650");win.configure(bg=CARD);win.transient(self);win.grab_set()
        tk.Label(win,text="Record shared overhead",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30)
        choices={f'{p["partner_code"]} — {p["name"]}':p["id"] for p in partners};paid=ttk.Combobox(form,values=list(choices),state="readonly");paid.set(next(iter(choices)) if choices else "")
        values=[("Batch code",generate_batch_code(self.db.setting("company_prefix"),self.db.setting("supply_date"))),("Expense date",date.today().isoformat()),("Category","Office"),("Description",""),("Amount (BDT)","0"),("Payment method","Cash"),("Reference number",""),("Voucher / file reference","")]
        entries={}
        for label,default in values:
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form);e.insert(0,default);e.pack(fill="x",pady=(4,9));entries[label]=e
        tk.Label(form,text="Paid by partner",bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");paid.pack(fill="x",pady=(4,9))
        def save():
            try:
                amount=float(entries["Amount (BDT)"].get());datetime.strptime(entries["Expense date"].get(),"%Y-%m-%d")
                if amount<=0 or not entries["Description"].get().strip():raise ValueError("Description and a positive amount are required")
                self.db.conn.execute("INSERT INTO overheads(batch_code,expense_date,category,description,amount_bdt,paid_by_partner_id,payment_method,reference_no,voucher) VALUES(?,?,?,?,?,?,?,?,?)",(entries["Batch code"].get().strip(),entries["Expense date"].get(),entries["Category"].get().strip(),entries["Description"].get().strip(),amount,choices.get(paid.get()),entries["Payment method"].get().strip(),entries["Reference number"].get().strip(),entries["Voucher / file reference"].get().strip()));self.db.conn.commit();win.destroy();self.partner_ledger()
            except ValueError as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save overhead",save).pack(fill="x",pady=8)

    def customers(self):
        body=self.page("Customers","Manage customer contacts, delivery addresses, and delivery status.")
        top=tk.Frame(body,bg=BG);top.pack(fill="x",pady=(0,12));self.action_button(top,"+ Add customer",lambda:self.customer_dialog()).pack(side="right")
        wrap=self.card(body);wrap.pack(fill="both",expand=True);tree=ttk.Treeview(wrap,columns=("id","name","phone","address","status","created"),show="headings")
        for c,l in zip(("id","name","phone","address","status","created"),("ID","Customer","Phone","Delivery address","Status","Created")):tree.heading(c,text=l)
        for c,w in zip(("id","name","phone","address","status","created"),(55,180,140,360,110,110)):tree.column(c,width=w)
        for row in self.db.conn.execute("SELECT * FROM customers ORDER BY id DESC"):tree.insert("","end",values=(row["id"],row["name"],row["phone"],row["address"],row["delivery_status"],row["created_at"][:10]))
        tree.pack(fill="both",expand=True,padx=15,pady=15)

    def customer_dialog(self):
        win=tk.Toplevel(self);win.title("Add customer");win.geometry("520x500");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="New customer",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30)
        entries={}
        for label,default in [("Name",""),("Phone",""),("Delivery address",""),("Delivery status","Active"),("Notes","")]:
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form);e.insert(0,default);e.pack(fill="x",pady=(4,11));entries[label]=e
        def save():
            try:
                if not all(entries[x].get().strip() for x in ("Name","Phone","Delivery address")):raise ValueError("Name, phone, and delivery address are required")
                self.db.conn.execute("INSERT INTO customers(name,phone,address,delivery_status,notes) VALUES(?,?,?,?,?)",tuple(entries[x].get().strip() for x in ("Name","Phone","Delivery address","Delivery status","Notes")));self.db.conn.commit();win.destroy();self.customers()
            except (ValueError,sqlite3.IntegrityError) as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save customer",save).pack(fill="x",pady=10)

    def suppliers(self):
        body=self.page("Suppliers","Compare supplier contacts, quote history, and quality evaluations.")
        top=tk.Frame(body,bg=BG);top.pack(fill="x",pady=(0,12));self.action_button(top,"+ Add supplier",self.supplier_dialog).pack(side="right");self.action_button(top,"+ Record quote",self.supplier_quote_dialog,secondary=True).pack(side="right",padx=(0,8))
        wrap=self.card(body);wrap.pack(fill="both",expand=True);tree=ttk.Treeview(wrap,columns=("id","name","contact","phone","wechat","quality","price"),show="headings",height=8)
        for c,l in zip(("id","name","contact","phone","wechat","quality","price"),("ID","Supplier","Contact person","Phone","WeChat","Quality / 5","Last price RMB")):tree.heading(c,text=l)
        for c,w in zip(("id","name","contact","phone","wechat","quality","price"),(50,210,150,130,140,100,120)):tree.column(c,width=w)
        for row in self.db.conn.execute("SELECT * FROM suppliers ORDER BY quality_rating DESC,name"):tree.insert("","end",values=(row["id"],row["name"],row["contact_person"],row["phone"],row["wechat"],f'{row["quality_rating"]:.1f}',f'{row["last_price_rmb"]:,.2f}'))
        tree.pack(fill="both",expand=True,padx=15,pady=(15,8))
        tk.Label(wrap,text="Recent supplier quotes",bg=CARD,fg=INK,font=("Segoe UI Semibold",12)).pack(anchor="w",padx=15,pady=(4,5))
        quotes=ttk.Treeview(wrap,columns=("date","supplier","product","price","notes"),show="headings",height=6)
        for c,l,w in (("date","Date",100),("supplier","Supplier",180),("product","Product",230),("price","Price RMB",110),("notes","Notes",260)):quotes.heading(c,text=l);quotes.column(c,width=w)
        for row in self.db.conn.execute("SELECT q.quoted_on,s.name,q.product_name,q.quoted_price_rmb,q.notes FROM supplier_prices q JOIN suppliers s ON s.id=q.supplier_id ORDER BY q.quoted_on DESC,q.id DESC LIMIT 100"):quotes.insert("","end",values=(row[0],row[1],row[2],f"{row[3]:,.2f}",row[4] or ""))
        quotes.pack(fill="both",expand=True,padx=15,pady=(0,15))

    def supplier_dialog(self):
        win=tk.Toplevel(self);win.title("Add supplier");win.geometry("540x680");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="New supplier",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30);entries={}
        for label,default in [("Supplier name",""),("Contact person",""),("Phone",""),("WeChat",""),("Address",""),("Quality rating (0–5)","0"),("Last price (RMB)","0"),("Notes","")]:
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form);e.insert(0,default);e.pack(fill="x",pady=(4,9));entries[label]=e
        def save():
            try:
                rating=float(entries["Quality rating (0–5)"].get());price=float(entries["Last price (RMB)"].get())
                if not 0<=rating<=5 or not entries["Supplier name"].get().strip():raise ValueError("Supplier name is required and rating must be between 0 and 5")
                self.db.conn.execute("INSERT INTO suppliers(name,contact_person,phone,wechat,address,quality_rating,last_price_rmb,notes) VALUES(?,?,?,?,?,?,?,?)",(entries["Supplier name"].get().strip(),entries["Contact person"].get().strip(),entries["Phone"].get().strip(),entries["WeChat"].get().strip(),entries["Address"].get().strip(),rating,price,entries["Notes"].get().strip()));self.db.conn.commit();win.destroy();self.suppliers()
            except (ValueError,sqlite3.IntegrityError) as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save supplier",save).pack(fill="x",pady=10)

    def supplier_quote_dialog(self):
        suppliers=self.db.conn.execute("SELECT id,name FROM suppliers ORDER BY name").fetchall()
        if not suppliers:return messagebox.showwarning("Supplier required","Add a supplier before recording a quote.")
        win=tk.Toplevel(self);win.title("Record supplier quote");win.geometry("520x500");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="Record supplier quote",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30)
        supplier_map={s["name"]:s["id"] for s in suppliers};supplier=ttk.Combobox(form,values=list(supplier_map),state="readonly");supplier.set(next(iter(supplier_map)));product=ttk.Entry(form);price=ttk.Entry(form);price.insert(0,"0");quoted=ttk.Entry(form);quoted.insert(0,date.today().isoformat());notes=ttk.Entry(form)
        for label,widget in (("Supplier",supplier),("Product",product),("Quoted price (RMB)",price),("Quote date",quoted),("Notes",notes)):tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");widget.pack(fill="x",pady=(4,11))
        def save():
            try:
                amount=float(price.get());datetime.strptime(quoted.get(),"%Y-%m-%d")
                if amount<0 or not product.get().strip():raise ValueError("Product is required and price cannot be negative")
                sid=supplier_map[supplier.get()];self.db.conn.execute("INSERT INTO supplier_prices(supplier_id,product_name,quoted_price_rmb,quoted_on,notes) VALUES(?,?,?,?,?)",(sid,product.get().strip(),amount,quoted.get(),notes.get().strip()));self.db.conn.execute("UPDATE suppliers SET last_price_rmb=? WHERE id=?",(amount,sid));self.db.conn.commit();win.destroy();self.suppliers()
            except ValueError as exc:messagebox.showerror("Cannot save",str(exc),parent=win)
        self.action_button(form,"Save quote",save).pack(fill="x",pady=10)

    def sales(self):
        body=self.page("Sales & invoices","Create retail or wholesale sales, deduct stock, and export branded invoices.")
        top=tk.Frame(body,bg=BG);top.pack(fill="x",pady=(0,12));self.action_button(top,"+ New sale",self.sale_dialog).pack(side="right")
        wrap=self.card(body);wrap.pack(fill="both",expand=True);tree=ttk.Treeview(wrap,columns=("id","invoice","date","customer","type","status","total","paid"),show="headings")
        for c,l in zip(("id","invoice","date","customer","type","status","total","paid"),("ID","Invoice","Date","Customer","Type","Delivery","Total BDT","Paid BDT")):tree.heading(c,text=l)
        for c,w in zip(("id","invoice","date","customer","type","status","total","paid"),(50,190,100,170,100,105,120,120)):tree.column(c,width=w)
        rows=self.db.conn.execute("SELECT s.*,c.name customer_name FROM sales s LEFT JOIN customers c ON c.id=s.customer_id ORDER BY s.id DESC").fetchall()
        for row in rows:tree.insert("","end",values=(row["id"],row["invoice_no"],row["sale_date"],row["customer_name"] or "Walk-in",row["sale_type"],row["delivery_status"],f'{row["total_bdt"]:,.2f}',f'{row["paid_bdt"]:,.2f}'))
        tree.pack(fill="both",expand=True,padx=15,pady=(15,8))
        def invoice():
            selected=tree.selection()
            if not selected:return messagebox.showwarning("Select a sale","Select a sale to export its invoice.")
            self.export_invoice_pdf(int(tree.item(selected[0],"values")[0]))
        self.action_button(wrap,"Export selected invoice PDF",invoice,secondary=True).pack(anchor="e",padx=15,pady=(0,15))

    def sale_dialog(self):
        customers=self.db.conn.execute("SELECT * FROM customers ORDER BY name").fetchall();products=self.db.products()
        if not customers:return messagebox.showwarning("Customer required","Add a customer before creating a sale.")
        if not products:return messagebox.showwarning("Product required","Add a product before creating a sale.")
        win=tk.Toplevel(self);win.title("New sale");win.geometry("560x600");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="Create sale",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30)
        customer_map={f'{c["name"]} — {c["phone"]}':c["id"] for c in customers};product_map={f'{p["code"]} — {p["name"]} (stock {p["quantity"]})':p for p in products}
        customer=ttk.Combobox(form,values=list(customer_map),state="readonly");customer.set(next(iter(customer_map)));product=ttk.Combobox(form,values=list(product_map),state="readonly");product.set(next(iter(product_map)));sale_type=ttk.Combobox(form,values=("Retail","Wholesale"),state="readonly");sale_type.set("Retail");qty=ttk.Entry(form);qty.insert(0,"1");price=ttk.Entry(form);delivery=ttk.Combobox(form,values=("Pending","Packed","Shipped","Delivered"),state="readonly");delivery.set("Pending")
        widgets=[("Customer",customer),("Product",product),("Sale type",sale_type),("Quantity",qty),("Unit price (BDT)",price),("Delivery status",delivery)]
        for label,widget in widgets:tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");widget.pack(fill="x",pady=(4,12))
        def refresh_price(*_):
            p=product_map[product.get()];all_products,_,rate,retail,wholesale,sourcing=self.context();total=sum(float(x["purchase_rmb"]) for x in all_products);c=calculate_cost(p,total,sourcing,rate,retail,wholesale);value=c.retail_bdt if sale_type.get()=="Retail" else c.wholesale_bdt;price.delete(0,"end");price.insert(0,f"{value:.2f}")
        product.bind("<<ComboboxSelected>>",refresh_price);sale_type.bind("<<ComboboxSelected>>",refresh_price);refresh_price()
        def save():
            try:
                sale_id=self.db.create_sale(customer_map[customer.get()],product_map[product.get()]["id"],int(qty.get()),sale_type.get(),float(price.get()),delivery.get(),self.user["id"]);win.destroy();self.sales();self.export_invoice_pdf(sale_id)
            except ValueError as exc:messagebox.showerror("Cannot create sale",str(exc),parent=win)
        self.action_button(form,"Create sale and invoice",save).pack(fill="x",pady=10)

    def export_invoice_pdf(self, sale_id: int):
        path=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF invoice","*.pdf")],initialfile=f"CBDS-invoice-{sale_id}.pdf")
        if not path:return
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image as PDFImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        sale=self.db.conn.execute("SELECT s.*,c.name customer_name,c.phone,c.address FROM sales s LEFT JOIN customers c ON c.id=s.customer_id WHERE s.id=?",(sale_id,)).fetchone();items=self.db.conn.execute("SELECT si.*,p.code,p.name FROM sale_items si JOIN products p ON p.id=si.product_id WHERE si.sale_id=?",(sale_id,)).fetchall()
        def brand(canvas,doc):
            canvas.saveState();canvas.setFillAlpha(.06);canvas.drawImage(str(resource_path("assets/cbds_watermark.png")),40*mm,85*mm,width=130*mm,height=90*mm,mask="auto",preserveAspectRatio=True);canvas.setFillAlpha(1);canvas.setFont("Helvetica",8);canvas.drawCentredString(105*mm,10*mm,"China BD Store • Value in Every Detail • From China to your home");canvas.restoreState()
        doc=SimpleDocTemplate(path,pagesize=A4,rightMargin=18*mm,leftMargin=18*mm,topMargin=12*mm,bottomMargin=18*mm);styles=getSampleStyleSheet();story=[PDFImage(str(resource_path("assets/cbds_logo.png")),width=55*mm,height=28*mm),Paragraph("<b>SALES INVOICE</b>",styles["Title"]),Paragraph(f"Invoice: {sale['invoice_no']}<br/>Date: {sale['sale_date']}<br/>Type: {sale['sale_type']}",styles["Normal"]),Spacer(1,5*mm),Paragraph(f"<b>Bill to:</b> {sale['customer_name'] or 'Walk-in'}<br/>{sale['phone'] or ''}<br/>{sale['address'] or ''}",styles["Normal"]),Spacer(1,6*mm)]
        data=[["Code","Product","Qty","Unit price BDT","Total BDT"]]+[[i["code"],i["name"],i["quantity"],f'{i["unit_price_bdt"]:,.2f}',f'{i["quantity"]*i["unit_price_bdt"]:,.2f}'] for i in items]+[["","","","Grand total",f'{sale["total_bdt"]:,.2f}']]
        table=Table(data,colWidths=[28*mm,70*mm,18*mm,30*mm,30*mm]);table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0B684F")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-2),.3,colors.HexColor("#D1D5DB")),("ALIGN",(2,1),(-1,-1),"RIGHT"),("FONTNAME",(3,-1),(-1,-1),"Helvetica-Bold"),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7)]));story.extend([table,Spacer(1,12*mm),Paragraph("Thank you for choosing China BD Store.",styles["Normal"])]);doc.build(story,onFirstPage=brand,onLaterPages=brand);messagebox.showinfo("Invoice exported",f"Watermarked invoice saved to:\n{path}")

    def reports(self):
        body=self.page("Business reports","Profitability, sales, partner settlement, customer, and supplier reporting.")
        bar=tk.Frame(body,bg=BG);bar.pack(fill="x",pady=(0,12))
        self.action_button(bar,"Export PDF",self.export_pdf,secondary=True).pack(side="right")
        self.action_button(bar,"Export Excel",self.export_xlsx,secondary=True).pack(side="right",padx=(0,8))
        totals=tk.Frame(body,bg=BG);totals.pack(fill="x",pady=(0,12));month=date.today().strftime("%Y-%m");monthly=float(self.db.conn.execute("SELECT COALESCE(SUM(total_bdt),0) FROM sales WHERE substr(sale_date,1,7)=?",(month,)).fetchone()[0]);weekly=float(self.db.conn.execute("SELECT COALESCE(SUM(total_bdt),0) FROM sales WHERE date(sale_date)>=date('now','-6 days')").fetchone()[0]);all_sales=float(self.db.conn.execute("SELECT COALESCE(SUM(total_bdt),0) FROM sales").fetchone()[0]);overhead=float(self.db.conn.execute("SELECT COALESCE(SUM(amount_bdt),0) FROM overheads").fetchone()[0])
        for label,value,color in (("Sales last 7 days",money(weekly),GREEN),("Sales this month",money(monthly),PRIMARY),("All recorded sales",money(all_sales),ORANGE),("Shared overheads",money(overhead),RED)):
            c=self.card(totals);c.pack(side="left",fill="x",expand=True,padx=(0,12));tk.Frame(c,bg=color,height=3).pack(fill="x");tk.Label(c,text=label,bg=CARD,fg=MUTED).pack(anchor="w",padx=16,pady=(13,4));tk.Label(c,text=value,bg=CARD,fg=INK,font=("Segoe UI Semibold",16)).pack(anchor="w",padx=16,pady=(0,13))
        wrap=self.card(body);wrap.pack(fill="both",expand=True)
        cols=("code","name","purchase","ship","allocated","landed","retail","wholesale","margin")
        tree=ttk.Treeview(wrap,columns=cols,show="headings")
        for c,l in zip(cols,("Code","Product","Purchase RMB","Ship RMB","Sourcing RMB","Landed BDT","Retail BDT","Wholesale BDT","Retail profit")):tree.heading(c,text=l)
        for c,w in zip(cols,(105,210,100,90,105,110,110,120,105)):tree.column(c,width=w,anchor="e" if c not in ("code","name") else "w")
        products,costs,_,_,_,_=self.context()
        for p,c in zip(products,costs):tree.insert("","end",values=(p["code"],p["name"],f"{c.purchase_rmb:,.2f}",f"{c.shipping_rmb:,.2f}",f"{c.sourcing_rmb:,.2f}",f"{c.total_bdt:,.2f}",f"{c.retail_bdt:,.2f}",f"{c.wholesale_bdt:,.2f}",f"{c.retail_bdt-c.total_bdt:,.2f}"))
        tree.pack(fill="both",expand=True,padx=15,pady=15)

    def report_rows(self):
        products,costs,_,_,_,_=self.context()
        return [[p["code"],p["name"],p["category"],p["quantity"],c.purchase_rmb,c.shipping_rmb,c.sourcing_rmb,c.total_bdt,c.retail_bdt,c.wholesale_bdt] for p,c in zip(products,costs)]

    def export_xlsx(self):
        path=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel workbook","*.xlsx")],initialfile=f"CBDS-profitability-{date.today()}.xlsx")
        if not path:return
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            wb=Workbook();ws=wb.active;ws.title="Profitability"
            ws.merge_cells("A1:J1");ws["A1"]="CHINA BD STORE — PROFITABILITY REPORT";ws["A1"].font=Font(size=18,bold=True,color="FFFFFF");ws["A1"].fill=PatternFill("solid",fgColor="0B684F");ws["A1"].alignment=Alignment(horizontal="center");ws.row_dimensions[1].height=32
            ws.merge_cells("A2:J2");ws["A2"]=f"Official CBDS report • Generated by {APP_NAME} {APP_VERSION} • {datetime.now():%Y-%m-%d %H:%M}";ws["A2"].font=Font(italic=True,color="666666");ws["A2"].alignment=Alignment(horizontal="center")
            logo=XLImage(resource_path("assets/cbds_logo.png"));logo.width=210;logo.height=105;ws.add_image(logo,"A4");ws.row_dimensions[4].height=82
            headers=["Code","Product","Category","Quantity","Purchase RMB","Shipping RMB","Sourcing RMB","Landed BDT","Retail BDT","Wholesale BDT"]
            for col,value in enumerate(headers,1):
                cell=ws.cell(6,col,value);cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="B91C2D");cell.alignment=Alignment(horizontal="center")
            thin=Side(style="thin",color="E5E7EB")
            for r_idx,row in enumerate(self.report_rows(),7):
                for c_idx,value in enumerate(row,1):
                    cell=ws.cell(r_idx,c_idx,value);cell.border=Border(bottom=thin);cell.fill=PatternFill("solid",fgColor="F5FAF8" if r_idx%2 else "FFFFFF")
                    if c_idx>=4:cell.number_format='#,##0.00'
            widths=[16,30,18,12,16,16,16,16,16,18]
            for i,width in enumerate(widths,1):ws.column_dimensions[chr(64+i)].width=width
            ws.freeze_panes="A7";ws.auto_filter.ref=f"A6:J{6+len(self.report_rows())}";ws.sheet_view.showGridLines=False
            ws.oddFooter.center.text="CBDS • China BD Store • Value in Every Detail • From China to your home"
            ws.oddFooter.center.size=9;ws.oddFooter.center.color="777777"
            ws.oddHeader.right.text="CBDS"
            customers=wb.create_sheet("Customers");customers.append(["Name","Phone","Delivery address","Status","Notes"])
            for row in self.db.conn.execute("SELECT name,phone,address,delivery_status,notes FROM customers ORDER BY name"):customers.append(list(row))
            suppliers=wb.create_sheet("Suppliers");suppliers.append(["Supplier","Contact person","Phone","WeChat","Address","Quality / 5","Last price RMB","Notes"])
            for row in self.db.conn.execute("SELECT name,contact_person,phone,wechat,address,quality_rating,last_price_rmb,notes FROM suppliers ORDER BY name"):suppliers.append(list(row))
            partners=wb.create_sheet("Partner Settlement");partners.append(["Partner ID","Partner","Investment BDT","Overhead paid","Equal overhead share","Settlement balance"])
            for p,invested,paid,share,balance in self.db.partner_summary():partners.append([p["partner_code"],p["name"],invested,paid,share,balance])
            sales=wb.create_sheet("Sales");sales.append(["Invoice","Date","Customer","Type","Delivery status","Total BDT","Paid BDT"])
            for row in self.db.conn.execute("SELECT s.invoice_no,s.sale_date,c.name,s.sale_type,s.delivery_status,s.total_bdt,s.paid_bdt FROM sales s LEFT JOIN customers c ON c.id=s.customer_id ORDER BY s.sale_date"):sales.append(list(row))
            for sheet in (customers,suppliers,partners,sales):
                sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions;sheet.sheet_view.showGridLines=False
                sheet.oddHeader.right.text="CBDS • CHINA BD STORE"
                sheet.oddFooter.center.text="CBDS • Official export • Value in Every Detail"
                for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="0B684F")
                for col in sheet.columns:sheet.column_dimensions[col[0].column_letter].width=min(42,max(12,max(len(str(c.value or "")) for c in col)+2))
            wb.save(path);messagebox.showinfo("Export complete",f"Branded Excel report saved to:\n{path}")
        except Exception as exc:messagebox.showerror("Export failed",str(exc))

    def export_pdf(self):
        path=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF document","*.pdf")],initialfile=f"CBDS-profitability-{date.today()}.pdf")
        if not path:return
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Image as PDFImage, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
            def branded_page(canvas,doc):
                canvas.saveState();canvas.setFillAlpha(.07);canvas.drawImage(str(resource_path("assets/cbds_watermark.png")),70*mm,45*mm,width=160*mm,height=80*mm,mask="auto",preserveAspectRatio=True,anchor="c")
                canvas.setFillAlpha(1);canvas.setFont("Helvetica",8);canvas.setFillColor(colors.HexColor("#6B7280"));canvas.drawCentredString(148.5*mm,9*mm,"CBDS • China BD Store • Value in Every Detail • From China to your home");canvas.restoreState()
            doc=SimpleDocTemplate(path,pagesize=landscape(A4),rightMargin=12*mm,leftMargin=12*mm,topMargin=10*mm,bottomMargin=17*mm)
            styles=getSampleStyleSheet();story=[PDFImage(str(resource_path("assets/cbds_logo.png")),width=60*mm,height=30*mm),Paragraph("<b>Product Profitability Report</b>",styles["Title"]),Paragraph(f"Generated by {APP_NAME} {APP_VERSION} on {datetime.now():%Y-%m-%d %H:%M}",styles["Normal"]),Spacer(1,5*mm)]
            headers=["Code","Product","Category","Qty","Buy RMB","Ship RMB","Source RMB","Landed BDT","Retail BDT","Wholesale BDT"]
            data=[headers]+[[str(v) if i<4 else f"{float(v):,.2f}" for i,v in enumerate(row)] for row in self.report_rows()]
            table=Table(data,repeatRows=1,colWidths=[22*mm,43*mm,27*mm,12*mm,20*mm,20*mm,22*mm,24*mm,24*mm,27*mm])
            table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0B684F")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.2),("GRID",(0,0),(-1,-1),.25,colors.HexColor("#D1D5DB")),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F3F8F6")]),("ALIGN",(3,1),(-1,-1),"RIGHT"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
            story.append(table);doc.build(story,onFirstPage=branded_page,onLaterPages=branded_page);messagebox.showinfo("Export complete",f"Watermarked PDF report saved to:\n{path}")
        except Exception as exc:messagebox.showerror("Export failed",str(exc))

    def settings(self):
        body=self.page("Settings","Control global exchange, margin, sourcing, and supply variables.")
        form=self.card(body);form.pack(anchor="nw",fill="x",ipadx=20,ipady=10)
        tk.Label(form,text="Business calculation defaults",bg=CARD,fg=INK,font=("Segoe UI Semibold",14)).grid(row=0,column=0,columnspan=2,sticky="w",padx=24,pady=(18,20))
        specs=[("RMB exchange rate (BDT)","rmb_rate"),("Retail margin (%)","retail_margin"),("Wholesale margin (%)","wholesale_margin"),("Total sourcing cost (RMB)","sourcing_cost_rmb"),("Company code prefix","company_prefix"),("Supply date (YYYY-MM-DD)","supply_date")]
        entries={}
        for i,(label,key) in enumerate(specs,1):
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).grid(row=i,column=0,sticky="w",padx=24,pady=9)
            e=ttk.Entry(form,width=34);e.grid(row=i,column=1,sticky="ew",padx=24,pady=9);e.insert(0,self.db.setting(key));entries[key]=e
        form.columnconfigure(1,weight=1)
        def save():
            try:
                for key in ("rmb_rate","retail_margin","wholesale_margin","sourcing_cost_rmb"):float(entries[key].get())
                datetime.strptime(entries["supply_date"].get(),"%Y-%m-%d")
                if not entries["company_prefix"].get().strip():raise ValueError("Company prefix is required")
                self.db.set_settings({k:e.get().strip() for k,e in entries.items()});messagebox.showinfo("Settings saved","All calculations have been updated.")
            except ValueError as exc:messagebox.showerror("Invalid settings",str(exc))
        self.action_button(form,"Save and recalculate",save).grid(row=len(specs)+1,column=1,sticky="e",padx=24,pady=(18,22))
        safety=self.card(body);safety.pack(anchor="nw",fill="x",pady=(16,0),ipadx=20,ipady=12);tk.Label(safety,text="Security and data safety",bg=CARD,fg=INK,font=("Segoe UI Semibold",14)).pack(anchor="w",padx=24,pady=(15,4));tk.Label(safety,text="Create a transactionally consistent database backup or add an authorized user.",bg=CARD,fg=MUTED).pack(anchor="w",padx=24,pady=(0,12));buttons=tk.Frame(safety,bg=CARD);buttons.pack(fill="x",padx=24,pady=(0,15));self.action_button(buttons,"Create backup",self.create_backup,secondary=True).pack(side="left");self.action_button(buttons,"Add user",self.user_dialog,secondary=True).pack(side="left",padx=8)

    def create_backup(self):
        path=filedialog.asksaveasfilename(defaultextension=".db",filetypes=[("CBDs database backup","*.db")],initialfile=f"CBDs-backup-{datetime.now():%Y%m%d-%H%M}.db")
        if path:self.db.backup(path);messagebox.showinfo("Backup complete",f"Encrypted credentials and business data were backed up to:\n{path}")

    def user_dialog(self):
        win=tk.Toplevel(self);win.title("Add authorized user");win.geometry("500x470");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="Add authorized user",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30);entries={}
        for label in ("Full name","Username","Password"):
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form,show="●" if label=="Password" else "");e.pack(fill="x",pady=(4,11));entries[label]=e
        role=ttk.Combobox(form,values=("partner","admin"),state="readonly");role.set("partner");tk.Label(form,text="Role",bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");role.pack(fill="x",pady=(4,11))
        def save():
            try:
                password=entries["Password"].get()
                if len(password)<10 or not entries["Full name"].get().strip() or not entries["Username"].get().strip():raise ValueError("Name and username are required; password must be at least 10 characters")
                salt,digest=hash_password(password);self.db.conn.execute("INSERT INTO users(name,username,role,salt,password_hash) VALUES(?,?,?,?,?)",(entries["Full name"].get().strip(),entries["Username"].get().strip(),role.get(),salt,digest));self.db.conn.commit();win.destroy();messagebox.showinfo("User created","The authorized user account was created.")
            except (ValueError,sqlite3.IntegrityError) as exc:messagebox.showerror("Cannot create user",str(exc),parent=win)
        self.action_button(form,"Create user",save).pack(fill="x",pady=10)

    def password_dialog(self):
        win=tk.Toplevel(self);win.title("Change password");win.geometry("500x430");win.configure(bg=CARD);win.transient(self);win.grab_set();tk.Label(win,text="Change your password",bg=CARD,fg=INK,font=("Segoe UI Semibold",19)).pack(anchor="w",padx=30,pady=(24,16));form=tk.Frame(win,bg=CARD);form.pack(fill="both",expand=True,padx=30);entries={}
        for label in ("Current password","New password","Confirm new password"):
            tk.Label(form,text=label,bg=CARD,fg=INK,font=("Segoe UI Semibold",9)).pack(anchor="w");e=ttk.Entry(form,show="●");e.pack(fill="x",pady=(4,11));entries[label]=e
        def save():
            try:
                if not self.db.authenticate(self.user["username"],entries["Current password"].get()):raise ValueError("Current password is incorrect")
                new=entries["New password"].get()
                if len(new)<10:raise ValueError("New password must be at least 10 characters")
                if new!=entries["Confirm new password"].get():raise ValueError("New passwords do not match")
                salt,digest=hash_password(new);self.db.conn.execute("UPDATE users SET salt=?,password_hash=? WHERE id=?",(salt,digest,self.user["id"]));self.db.conn.commit();win.destroy();messagebox.showinfo("Password changed","Your password was changed successfully.")
            except ValueError as exc:messagebox.showerror("Cannot change password",str(exc),parent=win)
        self.action_button(form,"Change password",save).pack(fill="x",pady=10)


if __name__ == "__main__":
    App().mainloop()
